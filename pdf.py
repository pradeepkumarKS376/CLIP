import openpyxl as openpyxl
import pandas as pd
import PyPDF2
import re

from xlsxwriter import Workbook

pdf_path = r"D:\Check\990702118000000000.PDF"
newFileName = r"D:\Check\New Text Document.pdf"
search_term = 'of'
rotation = 90
getPagecount = 0
splits = [1,2]
result_list = []
result = ""
columnname = ""
text=""
text1=""
result_list1 = []
def SearchText(pdf_path,search_term,text1):
    Word = ""
    EColon = ""
    MobNo = ""
    Colon = ""
    Email = ""
    pdfFileObj = open(pdf_path, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    for page_number in range(0, pdfReader.numPages):
        pageObj = pdfReader.getPage(page_number)
        page_content = pageObj.extractText()
        result = {"page": page_number,"content": page_content}.__str__().lower()
        Temp1 = re.findall(r'[a-zA-Z]\w+', result)
        Temp2 = re.findall('(\w+(?: \w+)*):', result)
        Temp3 = re.findall(r'([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+', result)
        Temp4 = re.findall(r'(\w+):+', result)
        Temp5 = re.findall(r'^.+:(.+)$', result)
        Word = Word + str(Temp1).replace('[','').replace(']','').replace('\\','').lower()
        Colon = Colon + str(Temp2).replace('[','').replace(']','').replace('\\','').lower()
        Email = Email + str(Temp3).replace('[', '').replace(']', '').replace('\\', '').lower()
        MobNo = MobNo + str(Temp4).replace('[','').replace(']','').replace('\\','').lower()
        EColon = EColon + str(Temp5).replace('[','').replace(']','').replace('\\','').lower()
        Temp1 = ""
        Temp2 = ""
        Temp3 = ""
        Temp4 = ""
        Temp5 = ""
    result_list.append(Word)
    result_list.append(Email)
    result_list.append(Colon)
    result_list.append(MobNo)
    result_list.append(EColon)
    df2 = pd.DataFrame(result_list,columns=['Sheet1'])

    for page_number in range(0, pdfReader.numPages):
        text = df2.iloc[page_number].iloc[0]
        text = text.replace("'", "")
        keywords = text.split(",")
        SheetName = 'keywords'+str(page_number)
        df = pd.DataFrame(list(set(keywords)), columns=['keywords'])
        df['number_of_times_word_appeared'] = df['keywords'].apply(lambda x: weightage(x, text))
        df['tf'] = df['keywords'].apply(lambda x: weightage(x, text))
        writer = pd.ExcelWriter('example.xlsx', engine='xlsxwriter')
        #wb = openpyxl.Workbook()
        #sheet = wb.create_sheet(SheetName, page_number)
        #print(wb.sheetnames)
        #Inc = 1
        #wb.active = Inc
        #print(wb.active)
        #print(SheetName)
        #Inc = Inc+1
        #wb.save("pandas_openpyxl.xlsx")
        df.to_excel(writer, sheet_name=SheetName)
        #wb.save('example.xlsx')
        writer.save()
    pdfFileObj.close()


def weightage(word, text, number_of_documents=1):
    word_list = re.findall(word, text)
    number_of_times_word_appeared = len(word_list)
    return number_of_times_word_appeared

SearchText(pdf_path,search_term,text1)